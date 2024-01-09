from openpyxl import load_workbook
import re
from pathlib import Path

print('hello world')
def getdownloadpath():
    downloads_path = str(Path.home() / "Downloads")
    return downloads_path
def excel_to_text():
    input_file_path = '/home/buzzadmin/Downloads/EPF_Regular_ECR.xlsx'
    if  input_file_path is not None:        
        input_file_path_split = input_file_path.split('.')[0]
        output_file_path= input_file_path_split +'.txt'    
        headers = ['UAN','MEMBER_NAME','GROSS_WAGES','EPF_WAGES','EPS_WAGES','EDLI_WAGES','EPF_CONTRI_REMITTED 12%','EPS_CONTRI_REMITTED 8.33%','EPF_EPS_DIFF_REMITTED 3.67%','NCP_DAYS','REFUND_OF_ADVANCES']

        # Load the workbook
        workbook = load_workbook(input_file_path, data_only=True)
        sheet = workbook.active
        # Open the output file in write mode
        with open(output_file_path, 'w') as output_file:
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is not None:  # Check if 'UAN' column is not null and not equal to 101953167960
                    row_new = [ str(cell) if cell is not None else '' for cell in row]
                    row_data = [str(int(float(test))) if re.match("^\d+?\.\d+?$",test) else str(test) for test in row_new]
                    filtered_row = [row_data[i] for i in range(len(headers))]  
                    output_file.write('#~#'.join(filtered_row) + '\n')
        print("Conversion to text file done successfully.")    
    return  output_file_path
    